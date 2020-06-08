import openpyxl
from openpyxl.utils import get_column_letter

wb = openpyxl.load_workbook('4年级.xlsx')

wb.save('new.xlsx')

print('length:' + str(len(wb.sheetnames)))
if len(wb.sheetnames) < 2:
    print('linjunjie')
    wb.create_sheet(title='4.1')
    wb.create_sheet(title='4.2')
    wb.create_sheet(title='4.3')
    wb.create_sheet(title='4.4')
    wb.create_sheet(title='4.5')
    wb.create_sheet(title='4.6')
    wb.create_sheet(title='4.7')
    wb.create_sheet(title='4.8')
    wb.create_sheet(title='4.9')
    wb.create_sheet(title='4.10')
    wb.create_sheet(title='4.11')
    wb.create_sheet(title='4.12')
#
sheet = wb.active
if sheet.cell(row=1, column=1).value == '序号':
    sheet.delete_cols(idx=1, amount=6)
    print('周结论')
# sheet.delete_cols(idx=1, amount=6)

# 1.拷贝第一行表头 只需要加一次
for cellList in sheet.iter_rows(min_row=1, max_row=1, max_col=sheet.max_column + 1):
    # for cell in cellList:
    for grade in wb.sheetnames:
        wb[grade].append((cell.value for cell in cellList))

# 2.遍历是否参加课后服务的数据,并进行替换
for rowNum in range(2, sheet.max_row + 1):
    isOrNotJoinService = sheet.cell(row=rowNum, column=6).value

    if isOrNotJoinService == '是' or isOrNotJoinService == '否':
        break
    if str(isOrNotJoinService) == '1':
        sheet.cell(row=rowNum, column=6).value = '是'
    elif str(isOrNotJoinService == '2'):
        sheet.cell(row=rowNum, column=6).value = '否'

# 3.遍历每一行 找到对应班级 添加到相应表里面
for rowNum in range(2, sheet.max_row + 1):
    targetRowNum = 1
    gradeValue = str(sheet.cell(row=rowNum, column=4).value)
    classValue = str(sheet.cell(row=rowNum, column=5).value)
    gradeClassValue = gradeValue + '.' + classValue

    sheet.cell(row=rowNum, column=4).value = gradeClassValue
    for sheetName in wb.sheetnames:
        if sheetName == gradeClassValue:
            targetRowNum = rowNum

    if targetRowNum != 1:
        # 移动到对应sheet表
        for cellList in sheet.iter_rows(min_row=targetRowNum, max_row=targetRowNum, max_col=sheet.max_column + 1):
            # for cell in cellList:
            wb[gradeClassValue].append((cell.value for cell in cellList))

# 4.删除多余的列
for sheetName in wb.sheetnames:
    current_sheet = wb[sheetName]
    print(current_sheet.cell(row=1, column=5).value)
    if sheetName != 'Sheet1' and current_sheet.cell(row=1, column=5).value == '3、学生班级':
        wb[sheetName].delete_cols(idx=5, amount=1)

wb.save('new.xlsx')
