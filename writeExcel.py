import openpyxl
from openpyxl.utils import get_column_letter

# 1. 更改sheet
# wb = openpyxl.Workbook()
# sheet = wb.active

# print(sheet.title)
# # change the name of sheet
# sheet.title = 'Happy2020'
# print(wb.sheetnames)
#
# wb.save('example_copy.xlsx')
#
# wb = openpyxl.load_workbook('original.xlsx')
# wb.save('altered.xlsx')
#
# wb.create_sheet(index=0, title='First Sheet')
# wb.create_sheet(index=1, title='Middle Sheet')
# print(wb.sheetnames)
#
# wb.remove(wb['Middle Sheet'])
# print(wb.sheetnames)
#
# wb.save('temp1.xlsx')

# 2. 写入cell数据
# wb = openpyxl.Workbook()
# sheet = wb.active
#
# sheet['A1'] = 'Hello Python'
# print(sheet['A1'].value)
#
# ws1 = wb.create_sheet('range names')
# for row in range(1, 40):
#     ws1.append(range(12))
#
# ws2 = wb.create_sheet('list')
# rows = [
#     ['Number', 'Batch 1', 'Batch 2'],
#     [2, 40, 30],
#     [3, 40, 25],
#     [4, 50, 55],
#     [5, 50, 15],
#     [6, 60, 35],
#     [7, 60, 45],
# ]
#
# for row in rows:
#     ws2.append(row)
#
# ws3 = wb.create_sheet(title='Data')
# for row in range(5, 30):
#     for col in range(15, 54):
#         ws3.cell(column=col, row=row, value=get_column_letter(col))
#
# print(ws3['AA10'])
# wb.save('empty_book.xlsx')


# 3.找到对应cell进行数据修改


