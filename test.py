# import census2010
#
# print(census2010.allData['AK']['Bethel']['pop'])


from openpyxl import Workbook

wb = Workbook()
ws = wb.active

data = [
    ["Fruit", "Quantity"],
    ["Kiwi", 3],
    ["Grape", 15],
    ["Apple", 3],
    ["Peach", 3],
    ["Pomegranate", 3],
    ["Pear", 3],
    ["Tangerine", 3],
    ["Blueberry", 43],
    ["Mango", 13],
    ["Watermelon", 3],
    ["Blackberry", 3],
    ["Orange", 3],
    ["Raspberry", 3],
    ["Banana", 3]
]

for r in data:
    ws.append(r)

ws.auto_filter.ref = "A1:B15"
# ws.auto_filter.add_filter_column(0)
# ws.auto_filter.add_filter_column(0, ["Kiwi", "Apple", "Mango"])
ws.auto_filter.add_sort_condition("B2:B15")

wb.create_sheet(title='123')

for rowObj in ws.iter_rows():
    test = rowObj[ws.min_column - 1:ws.max_column]
    tempValue = (cell.value for cell in rowObj)
    print(list(tempValue)[1])
    # for value in list(tempValue)                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                    :
    # wb['123'].append(tempValue)




# wb.save("filtered.xlsx")